"""تحلیل هوشمند پیام‌های روبیکا — فاز ۷ سند.

پنج تابع مستقل — هر کدام یک Celery task مجزا دارد:

الف) extract_prices_from_messages  — GPT-4o-mini، هر شب ساعت ۲۲
ب)  detect_complaints_and_alert    — GPT-4o-mini، هر ۵ دقیقه
ج)  handle_conversation_reply      — GPT-4o-mini، on-demand (توسط listener)
د)  count_red_keywords             — pure SQL، بدون AI
ه)  generate_daily_summary         — DeepSeek (ollama/OpenAI-compat)، هر شب ساعت ۲۲

DeepSeek:
  اگر DEEPSEEK_BASE_URL تنظیم باشد → OpenAI-compat endpoint (معمولاً http://localhost:11434/v1)
  اگر نه → به gpt-4o-mini fallback می‌کند (با هشدار)

Excel قیمت:
  storage/price_reports/rubika/YYYY-MM-DD_{group_guid[:12]}.xlsx
  یک شیت: ستون‌های محصول، قیمت، واحد، منبع (گروه)، زمان
"""

from __future__ import annotations

import logging
import os
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from sqlalchemy import func
from sqlalchemy.orm import Session

from core_engine.models import RubikaAllowedGroup, RubikaGroupMessage

if TYPE_CHECKING:
    pass

logger = logging.getLogger("core_engine.services.rubika_ai_analyzer")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
PRICE_REPORTS_DIR = PROJECT_ROOT / "storage" / "price_reports" / "rubika"

_MINI_MODEL = "gpt-4o-mini"
_DEEPSEEK_MODEL = os.getenv("DEEPSEEK_MODEL", "deepseek-r1:8b")

# ─── LLM helpers ──────────────────────────────────────────────────────────────


async def _chat(
    *,
    system: str,
    user: str,
    model: str = _MINI_MODEL,
    max_tokens: int = 800,
    use_deepseek: bool = False,
) -> str:
    """یک پاسخ متنی از GPT-4o-mini یا DeepSeek (fallback) دریافت می‌کند."""
    from openai import AsyncOpenAI

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    base_url: str | None = None

    if use_deepseek:
        deepseek_url = os.getenv("DEEPSEEK_BASE_URL", "").strip()
        if deepseek_url:
            base_url = deepseek_url.rstrip("/") + "/v1"
            api_key = os.getenv("DEEPSEEK_API_KEY", "ollama").strip() or "ollama"
            model = _DEEPSEEK_MODEL
        else:
            logger.warning(
                "rubika_ai: DEEPSEEK_BASE_URL تنظیم نشده — به %s fallback", _MINI_MODEL
            )

    if not api_key:
        raise ValueError("rubika_ai: OPENAI_API_KEY تنظیم نشده — تحلیل AI ناموفق")

    client = AsyncOpenAI(api_key=api_key, base_url=base_url)
    response = await client.chat.completions.create(
        model=model,
        max_tokens=max_tokens,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        temperature=0.3,
    )
    return (response.choices[0].message.content or "").strip()


# ─── الف — Price Extractor ─────────────────────────────────────────────────────


_PRICE_SYSTEM = """تو یک دستیار تخصصی استخراج قیمت از پیام‌های بازار ایران هستی.
متن‌های فارسی را تحلیل کن. اگر قیمت یا اطلاعات محصولی یافتی، آن را به این فرمت JSON برگردان:
[{"product": "نام محصول", "price": "عدد یا بازه", "unit": "واحد (تومان/دلار/...)", "raw": "عبارت اصلی"}]
اگر هیچ قیمتی نیافتی: []
فقط JSON برگردان، بدون توضیح اضافه."""


def _write_price_excel(
    group_guid: str,
    group_name: str | None,
    target_date: date,
    price_rows: list[dict],
) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    PRICE_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    safe = group_guid.replace("/", "_")[:12]
    path = PRICE_REPORTS_DIR / f"{target_date.isoformat()}_{safe}.xlsx"

    if path.exists():
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "قیمت‌ها"
        headers = ["محصول", "قیمت", "واحد", "گروه", "زمان ثبت", "متن خام"]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="right")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 40

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row in price_rows:
        ws.append([
            row.get("product", ""),
            row.get("price", ""),
            row.get("unit", ""),
            group_name or group_guid,
            now_str,
            row.get("raw", ""),
        ])

    wb.save(str(path))
    return path


async def extract_prices_from_messages(
    db: Session, *, group_guid: str, target_date: date | None = None
) -> dict:
    """پیام‌های متنی+transcription+image_extracted_text روز مشخص را تحلیل می‌کند.

    Returns:
        dict: total_messages, price_rows_found, excel_path
    """
    d = target_date or date.today()
    start = datetime.combine(d, datetime.min.time())
    end = datetime.combine(d, datetime.max.time())

    rows = (
        db.query(RubikaGroupMessage)
        .filter(
            RubikaGroupMessage.group_guid == group_guid,
            RubikaGroupMessage.received_at.between(start, end),
        )
        .all()
    )

    if not rows:
        return {"total_messages": 0, "price_rows_found": 0, "excel_path": None}

    texts = []
    for r in rows:
        if r.message_text:
            texts.append(r.message_text)
        if r.transcription and r.transcription not in ("[پردازش ناموفق]", "[بدون اطلاعات تجاری]"):
            texts.append(r.transcription)
        if r.image_extracted_text and r.image_extracted_text not in ("[پردازش ناموفق]", "[بدون اطلاعات تجاری]"):
            texts.append(r.image_extracted_text)

    if not texts:
        return {"total_messages": len(rows), "price_rows_found": 0, "excel_path": None}

    combined = "\n".join(texts)
    # محدودیت context — GPT-4o-mini حدود ۱۲۸k token دارد؛ بریدن به ۸۰۰۰ کاراکتر ایمن است
    if len(combined) > 8000:
        combined = combined[:8000] + "\n...[ادامه کوتاه شده]"

    import json

    try:
        raw_response = await _chat(
            system=_PRICE_SYSTEM,
            user=combined,
            model=_MINI_MODEL,
            max_tokens=1500,
        )
        # پاک‌سازی ```json ... ```
        clean = raw_response.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        price_rows = json.loads(clean)
        if not isinstance(price_rows, list):
            price_rows = []
    except Exception:
        logger.exception("rubika_ai: price extraction failed for group %s", group_guid)
        return {"total_messages": len(rows), "price_rows_found": 0, "excel_path": None, "error": True}

    group = db.query(RubikaAllowedGroup).filter(
        RubikaAllowedGroup.group_guid == group_guid
    ).first()
    group_name = group.group_name if group else None

    excel_path: Path | None = None
    if price_rows:
        excel_path = _write_price_excel(group_guid, group_name, d, price_rows)
        logger.info(
            "rubika_ai: %d price rows → %s", len(price_rows), excel_path
        )

    return {
        "total_messages": len(rows),
        "price_rows_found": len(price_rows),
        "excel_path": str(excel_path) if excel_path else None,
    }


# ─── ب — Complaint Detector ────────────────────────────────────────────────────


_COMPLAINT_SYSTEM = """تو یک دستیار تشخیص شکایت در پیام‌های فارسی هستی.
یک پیام دریافت می‌کنی. تشخیص بده آیا این پیام حاوی شکایت، نارضایتی، کنسلی، بازگشت کالا، یا اعتراض است.
پاسخ فقط: {"complaint": true, "reason": "دلیل کوتاه فارسی"} یا {"complaint": false}"""


async def detect_complaints_and_alert(db: Session, *, message_id: int) -> bool:
    """یک پیام را تحلیل می‌کند. اگر شکایت بود True + لاگ هشدار برمی‌گرداند.

    نوتیف واقعی (فاز ۴ پنل): از طریق GET /rubika/groups/{id}/messages با
    is_reply_to_our_message فیلتر می‌شود — اینجا فقط ساخت‌یافته لاگ می‌شود.
    """
    row = db.query(RubikaGroupMessage).filter(RubikaGroupMessage.id == message_id).first()
    if not row:
        return False

    text = row.message_text or row.transcription or ""
    if not text.strip():
        row.ai_analyzed = True
        db.flush()
        return False

    import json

    try:
        raw = await _chat(
            system=_COMPLAINT_SYSTEM,
            user=text[:1000],
            model=_MINI_MODEL,
            max_tokens=100,
        )
        clean = raw.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        result = json.loads(clean)
        is_complaint = bool(result.get("complaint"))
        reason = result.get("reason", "")
    except Exception:
        logger.exception("rubika_ai: complaint detection failed for message %s", message_id)
        row.ai_analyzed = True
        db.flush()
        return False

    row.ai_analyzed = True
    db.flush()

    if is_complaint:
        logger.warning(
            "rubika_complaint_detected message_id=%s group=%s sender=%s reason=%s text_preview=%s",
            message_id,
            row.group_guid,
            row.sender_name,
            reason,
            text[:100],
        )
    return is_complaint


# ─── ج — Conversation Handler ──────────────────────────────────────────────────


_CONVERSATION_SYSTEM = """تو یک دستیار فروش فارسی‌زبان هستی که از طرف یک کسب‌وکار
با مشتریان روبیکا مکالمه می‌کنی. پاسخ‌ها باید:
- فارسی روان و طبیعی باشند
- کوتاه (حداکثر ۳ جمله)
- مودبانه و حرفه‌ای
اگر سوال تخصصی است که جواب نمی‌دانی بگو: "کارشناس ما در اسرع وقت پاسخ می‌دهند."
"""


async def handle_conversation_reply(
    db: Session, *, message_id: int
) -> str | None:
    """اگر پیام ریپلای به ما بود و conversation_mode فعال است، پاسخ هوشمند تولید می‌کند.

    Returns:
        متن پاسخ یا None اگر باید رد شود.
    """
    row = db.query(RubikaGroupMessage).filter(RubikaGroupMessage.id == message_id).first()
    if not row or not row.is_reply_to_our_message:
        return None

    group = db.query(RubikaAllowedGroup).filter(
        RubikaAllowedGroup.group_guid == row.group_guid
    ).first()
    if not group or not group.conversation_mode_enabled:
        return None

    # تاریخچه: ۵ پیام اخیر همین گروه (برای context)
    history_rows = (
        db.query(RubikaGroupMessage)
        .filter(
            RubikaGroupMessage.group_guid == row.group_guid,
            RubikaGroupMessage.id < message_id,
        )
        .order_by(RubikaGroupMessage.received_at.desc())
        .limit(5)
        .all()
    )
    history_rows.reverse()

    context_parts = []
    for h in history_rows:
        text = h.message_text or h.transcription or ""
        if text.strip():
            sender = h.sender_name or "مشتری"
            context_parts.append(f"{sender}: {text[:200]}")
    context_parts.append(f"{row.sender_name or 'مشتری'}: {row.message_text or row.transcription or ''}")
    context = "\n".join(context_parts)

    try:
        reply = await _chat(
            system=_CONVERSATION_SYSTEM,
            user=context[-2000:],  # آخر context اگر بلند بود
            model=_MINI_MODEL,
            max_tokens=200,
        )
        row.ai_analyzed = True
        db.flush()
        return reply.strip() or None
    except Exception:
        logger.exception("rubika_ai: conversation reply failed for message %s", message_id)
        row.ai_analyzed = True
        db.flush()
        return None


# ─── د — Red Keyword Counter (pure SQL) ───────────────────────────────────────


def count_red_keywords(db: Session, *, group_guid: str, target_date: date | None = None) -> dict:
    """شمارش پیام‌های دارای کلمه قرمز در یک گروه برای یک روز.

    Returns:
        dict: group_guid, date, total_red, total_messages
    """
    d = target_date or date.today()
    start = datetime.combine(d, datetime.min.time())
    end = datetime.combine(d, datetime.max.time())

    total_messages = (
        db.query(func.count(RubikaGroupMessage.id))
        .filter(
            RubikaGroupMessage.group_guid == group_guid,
            RubikaGroupMessage.received_at.between(start, end),
        )
        .scalar()
        or 0
    )

    total_red = (
        db.query(func.count(RubikaGroupMessage.id))
        .filter(
            RubikaGroupMessage.group_guid == group_guid,
            RubikaGroupMessage.received_at.between(start, end),
            RubikaGroupMessage.has_red_keyword.is_(True),
        )
        .scalar()
        or 0
    )

    return {
        "group_guid": group_guid,
        "date": d.isoformat(),
        "total_red": total_red,
        "total_messages": total_messages,
    }


# ─── ه — Daily Summarizer ─────────────────────────────────────────────────────


_SUMMARY_SYSTEM = """تو یک خلاصه‌نویس حرفه‌ای فارسی هستی.
متن مکالمات یک گروه تجاری را دریافت می‌کنی.
یک خلاصه ۵ تا ۱۰ جمله‌ای فارسی بنویس که:
- مهم‌ترین موضوعات بازار را پوشش دهد
- قیمت‌های کلیدی ذکر شده را خلاصه کند
- شکایات یا مسائل مهم را برجسته کند
فقط خلاصه را بنویس، بدون مقدمه."""


async def generate_daily_summary(
    db: Session, *, group_guid: str, target_date: date | None = None
) -> str:
    """خلاصه روزانه گروه با DeepSeek محلی (fallback: gpt-4o-mini).

    Returns:
        متن خلاصه فارسی ۵-۱۰ جمله‌ای.
    """
    d = target_date or date.today()
    start = datetime.combine(d, datetime.min.time())
    end = datetime.combine(d, datetime.max.time())

    rows = (
        db.query(RubikaGroupMessage)
        .filter(
            RubikaGroupMessage.group_guid == group_guid,
            RubikaGroupMessage.received_at.between(start, end),
        )
        .order_by(RubikaGroupMessage.received_at.asc())
        .all()
    )

    if not rows:
        return "امروز هیچ پیامی در این گروه ثبت نشده بود."

    lines = []
    for r in rows:
        text = r.message_text or r.transcription or r.image_extracted_text or ""
        if text.strip() and text not in ("[پردازش ناموفق]", "[بدون اطلاعات تجاری]"):
            sender = r.sender_name or "ناشناس"
            lines.append(f"{sender}: {text[:300]}")

    if not lines:
        return "پیام‌های امروز محتوای قابل خلاصه‌سازی نداشتند."

    full_text = "\n".join(lines)
    # محدودیت context برای DeepSeek محلی (۴k ایمن‌تر است)
    if len(full_text) > 4000:
        full_text = full_text[:4000] + "\n...[ادامه کوتاه شده]"

    try:
        summary = await _chat(
            system=_SUMMARY_SYSTEM,
            user=f"مکالمات گروه — تاریخ {d.isoformat()}:\n\n{full_text}",
            model=_DEEPSEEK_MODEL,
            max_tokens=600,
            use_deepseek=True,
        )
        group = db.query(RubikaAllowedGroup).filter(
            RubikaAllowedGroup.group_guid == group_guid
        ).first()
        group_name = (group.group_name if group else None) or group_guid

        logger.info(
            "rubika_ai: daily summary generated for group '%s' (%d msgs)",
            group_name, len(rows),
        )
        return summary or "خلاصه‌سازی پاسخ خالی برگرداند."
    except Exception:
        logger.exception("rubika_ai: daily summary failed for group %s", group_guid)
        return "[خلاصه‌سازی ناموفق بود]"
